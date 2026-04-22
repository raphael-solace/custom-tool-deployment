"""
Shared Excel manipulation tools for Solace Agent Mesh agents.

These tools intentionally keep the interface simple:
- inspect workbooks and preview rows
- update specific cells
- append rows to a worksheet

Complex payloads are passed as JSON strings so the tool schema remains stable
across agent runtimes.
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook
from openpyxl.workbook import Workbook


def _resolve_path(file_path: str) -> Path:
    path = Path(file_path).expanduser()
    if not path.exists():
        raise FileNotFoundError(f"Workbook not found: {path}")
    if path.suffix.lower() not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        raise ValueError(f"Unsupported workbook type: {path.suffix or 'unknown'}")
    return path


def _select_sheet(workbook: Workbook, sheet_name: Optional[str]) -> Any:
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"Worksheet '{sheet_name}' not found. Available sheets: {workbook.sheetnames}"
            )
        return workbook[sheet_name]
    return workbook[workbook.sheetnames[0]]


def _json_payload(payload: str, label: str) -> Any:
    try:
        return json.loads(payload)
    except json.JSONDecodeError as exc:
        raise ValueError(f"{label} must be valid JSON: {exc}") from exc


def _normalize_output_path(
    source_path: Path,
    output_path: Optional[str],
    suffix: str,
) -> Path:
    if output_path:
        return Path(output_path).expanduser()
    return source_path.with_name(f"{source_path.stem}{suffix}{source_path.suffix}")


def _coerce_cell_value(value: Any) -> Any:
    if isinstance(value, dict) and "value" in value:
        return value["value"]
    return value


def inspect_excel_workbook(
    file_path: str,
    sheet_name: Optional[str] = None,
    preview_rows: int = 5,
) -> Dict[str, Any]:
    """
    Return workbook structure and a small preview for agent reasoning.

    Args:
        file_path: Absolute or relative path to an Excel workbook.
        sheet_name: Optional worksheet name. Defaults to the first worksheet.
        preview_rows: Number of rows to preview, including the header row.
    """
    workbook_path = _resolve_path(file_path)
    workbook = load_workbook(workbook_path, data_only=True)
    worksheet = _select_sheet(workbook, sheet_name)

    rows = list(
        worksheet.iter_rows(
            min_row=1,
            max_row=max(1, preview_rows),
            values_only=True,
        )
    )
    header = list(rows[0]) if rows else []
    preview = [list(row) for row in rows[1:]] if len(rows) > 1 else []

    return {
        "status": "success",
        "file_path": str(workbook_path),
        "active_sheet": worksheet.title,
        "sheet_names": workbook.sheetnames,
        "max_row": worksheet.max_row,
        "max_column": worksheet.max_column,
        "header": header,
        "preview_rows": preview,
    }


def update_excel_cells(
    file_path: str,
    updates_json: str,
    sheet_name: Optional[str] = None,
    output_path: Optional[str] = None,
    create_copy: bool = True,
) -> Dict[str, Any]:
    """
    Update one or more cells in a workbook.

    Args:
        file_path: Path to the source workbook.
        updates_json: JSON array like
            [{"cell":"B2","value":"Approved"},{"cell":"C2","value":1250}]
        sheet_name: Optional worksheet name. Defaults to the first worksheet.
        output_path: Optional explicit destination path.
        create_copy: When true, write to a sibling copy instead of overwriting.
    """
    workbook_path = _resolve_path(file_path)
    updates = _json_payload(updates_json, "updates_json")
    if not isinstance(updates, list) or not updates:
        raise ValueError("updates_json must be a non-empty JSON array")

    workbook = load_workbook(workbook_path)
    worksheet = _select_sheet(workbook, sheet_name)

    applied_updates: List[Dict[str, Any]] = []
    for item in updates:
        if not isinstance(item, dict) or "cell" not in item:
            raise ValueError("Each update must be an object with at least a 'cell' field")
        cell_ref = str(item["cell"]).strip()
        if not cell_ref:
            raise ValueError("Cell reference cannot be empty")
        value = _coerce_cell_value(item.get("value"))
        worksheet[cell_ref] = value
        applied_updates.append({"cell": cell_ref, "value": value})

    destination = (
        _normalize_output_path(workbook_path, output_path, ".updated")
        if create_copy
        else Path(output_path).expanduser() if output_path else workbook_path
    )
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)

    return {
        "status": "success",
        "source_path": str(workbook_path),
        "output_path": str(destination),
        "sheet_name": worksheet.title,
        "updated_cells": applied_updates,
    }


def append_excel_rows(
    file_path: str,
    rows_json: str,
    sheet_name: Optional[str] = None,
    output_path: Optional[str] = None,
    create_copy: bool = True,
) -> Dict[str, Any]:
    """
    Append rows to a worksheet.

    Args:
        file_path: Path to the source workbook.
        rows_json: JSON array of arrays or objects. When objects are provided,
            keys are matched to the header row.
        sheet_name: Optional worksheet name. Defaults to the first worksheet.
        output_path: Optional explicit destination path.
        create_copy: When true, write to a sibling copy instead of overwriting.
    """
    workbook_path = _resolve_path(file_path)
    rows_payload = _json_payload(rows_json, "rows_json")
    if not isinstance(rows_payload, list) or not rows_payload:
        raise ValueError("rows_json must be a non-empty JSON array")

    workbook = load_workbook(workbook_path)
    worksheet = _select_sheet(workbook, sheet_name)

    header_cells = next(
        worksheet.iter_rows(min_row=1, max_row=1, values_only=True),
        tuple(),
    )
    headers = [str(value) if value is not None else "" for value in header_cells]

    appended_count = 0
    for row in rows_payload:
        if isinstance(row, list):
            worksheet.append(row)
            appended_count += 1
            continue

        if isinstance(row, dict):
            if not any(headers):
                raise ValueError(
                    "Cannot append object rows without a header row in the worksheet"
                )
            ordered_row = [row.get(header) for header in headers]
            worksheet.append(ordered_row)
            appended_count += 1
            continue

        raise ValueError("Each row must be either a JSON array or object")

    destination = (
        _normalize_output_path(workbook_path, output_path, ".updated")
        if create_copy
        else Path(output_path).expanduser() if output_path else workbook_path
    )
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)

    return {
        "status": "success",
        "source_path": str(workbook_path),
        "output_path": str(destination),
        "sheet_name": worksheet.title,
        "appended_rows": appended_count,
    }
